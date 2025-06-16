import openpyxl
import re


def sanitize_column_name(name):
    """
    Sanitiza o nome da coluna lido do XLSX para ser válido em SQL (PostgreSQL).
    Remove/substitui caracteres inválidos, remove espaços extras e converte para minúsculas.
    """
    if name is None:
        return None

    sanitized = str(name).strip()
    sanitized = sanitized.replace(' ', '_')
    sanitized = re.sub(r'[^\w]+', '', sanitized)
    sanitized = re.sub(r'_{2,}', '_', sanitized)
    sanitized = sanitized.strip('_')
    
    if sanitized and sanitized[0].isdigit():
        sanitized = '_' + sanitized

    sanitized = sanitized.lower()

    if not sanitized:

        return None
    return sanitized


def generate_create_table_sql_from_xlsx(file_path, sheet_name=None, table_name='microdados_ed_basica_2024',
                                        column_name_col_idx=1, column_type_col_idx=2, column_size_col_idx=3,
                                        header_row_idx=1):
    """
    Gera o script SQL CREATE TABLE lendo informações de um arquivo XLSX.

    Args:
        file_path (str): Caminho para o arquivo XLSX.
        sheet_name (str, optional): Nome da planilha a ser lida. Se None, usa a planilha ativa.
        table_name (str, optional): Nome desejado para a tabela no PostgreSQL.
        column_name_col_idx (int, optional): Índice da coluna (base 0) com os nomes das variáveis.
        column_type_col_idx (int, optional): Índice da coluna (base 0) com os tipos das variáveis.
        column_size_col_idx (int, optional): Índice da coluna (base 0) com o tamanho das variáveis.
        header_row_idx (int, optional): Índice da linha (base 1) que contém o cabeçalho, para pular.

    Returns:
        str: O script SQL CREATE TABLE gerado ou uma mensagem de erro.
    """

    TYPE_MAPPING = {
        'TIPO': 'TEXT',
        'NUM': 'INTEGER',
        'INTEIRO': 'INTEGER',
        'VARCHAR': 'VARCHAR',
        'TEXTO': 'TEXT',
        'TEXT': 'TEXT',
        'DATA': 'DATE',
        'DATE': 'DATE',
        'NUMÉRICO': 'NUMERIC',
        'NUMERIC': 'NUMERIC',
        'DECIMAL': 'DECIMAL',
        'BOOLEAN': 'BOOLEAN',
        'LOGICO': 'BOOLEAN',
        'BIGINT': 'BIGINT',
        'FLOAT': 'REAL',
        'DUPLO': 'DOUBLE PRECISION',

    }

    column_definitions = []

    try:

        workbook = openpyxl.load_workbook(file_path)

        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                return f"Erro: Planilha '{sheet_name}' não encontrada no arquivo."
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
            print(f"Usando a planilha ativa: '{sheet.title}'")

        for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
            row = sheet[row_idx]

            try:
                col_name_raw = row[column_name_col_idx].value
                col_type_raw = row[column_type_col_idx].value
                col_size_raw = None
                if column_size_col_idx < len(row):
                    col_size_raw = row[column_size_col_idx].value
            except IndexError:
                print(
                    f"Aviso: Pulando linha {row_idx} porque não tem colunas suficientes para nome/tipo.")
                continue

            col_name_sanitized = sanitize_column_name(col_name_raw)

            if not col_name_sanitized or col_type_raw is None or str(col_type_raw).strip() == '':
                print(f"Aviso: Pulando linha {row_idx} devido a nome ou tipo de coluna ausente/inválido (Nome Original: '{col_name_raw}', Tipo Original: '{col_type_raw}'). Nome Sanitizado: '{col_name_sanitized if col_name_sanitized else 'None'}'")
                continue

            col_type_str = str(col_type_raw).strip().upper()

            mapped_type = TYPE_MAPPING.get(col_type_str, 'TEXT')

            final_type = mapped_type

            col_definition = f'"{col_name_sanitized}" {final_type}'

            if col_size_raw is not None:
                final_type_upper = final_type.upper()
                if final_type_upper.startswith('VARCHAR') or final_type_upper.startswith('CHARACTER VARYING') or \
                   final_type_upper.startswith('CHAR') or final_type_upper.startswith('CHARACTER'):
                    try:
                        size_value = int(col_size_raw)
                        if size_value > 0:
                            col_definition += f'({size_value})'
                        else:
                            print(
                                f"Aviso: Tamanho '{col_size_raw}' na linha {row_idx} para '{col_name_sanitized}' ('{col_type_raw}' -> '{final_type}') é <= 0. Ignorando tamanho.")
                    except (ValueError, TypeError):
                        print(
                            f"Aviso: Tamanho '{col_size_raw}' na linha {row_idx} para '{col_name_sanitized}' ('{col_type_raw}' -> '{final_type}') não é um número inteiro válido para tipo texto. Ignorando tamanho.")

                elif final_type_upper.startswith('NUMERIC') or final_type_upper.startswith('DECIMAL'):
                    size_str = str(col_size_raw).strip().replace(
                        '(', '').replace(')', '')
                    if size_str:
                        parts = size_str.split(',')
                        try:
                            if len(parts) == 1:
                                precision = int(parts[0].strip())
                                if precision > 0:
                                    col_definition += f'({precision})'
                                else:
                                    print(
                                        f"Aviso: Precisão '{parts[0].strip()}' na linha {row_idx} para '{col_name_sanitized}' ('{col_type_raw}' -> '{final_type}') é <= 0. Ignorando precisão.")
                            elif len(parts) == 2:
                                precision = int(parts[0].strip())
                                scale = int(parts[1].strip())
                                if precision > 0 and scale >= 0:
                                    col_definition += f'({precision}, {scale})'
                                else:
                                    print(
                                        f"Aviso: Precisão ({precision}) ou Escala ({scale}) inválida na linha {row_idx} para '{col_name_sanitized}' ('{col_type_raw}' -> '{final_type}'). Ignorando tamanho/precisão.")
                            else:
                                print(
                                    f"Aviso: Formato de tamanho/precisão '{col_size_raw}' na linha {row_idx} para '{col_name_sanitized}' ('{col_type_raw}' -> '{final_type}') não é um formato válido (ex: '10' ou '10,2') para NUMERIC/DECIMAL. Ignorando tamanho.")
                        except (ValueError, TypeError):
                            print(
                                f"Aviso: Valores de tamanho/precisão '{col_size_raw}' na linha {row_idx} para '{col_name_sanitized}' ('{col_type_raw}' -> '{final_type}') não são números válidos para NUMERIC/DECIMAL. Ignorando tamanho.")

            column_definitions.append(col_definition)

    except FileNotFoundError:
        return f"Erro: Arquivo XLSX não encontrado em {file_path}"
    except Exception as e:
        return f"Ocorreu um erro inesperado ao ler o arquivo XLSX: {e}"

    if not column_definitions:
        return "-- Erro: Nenhuma definição de coluna válida encontrada no arquivo XLSX para gerar o SQL."

    # Gera o SQL CREATE TABLE
    sql = f'CREATE TABLE "{table_name}" (\n'
    sql += "    " + ",\n    ".join(column_definitions)
    sql += "\n);"
    sql += "\n\n-- Observações:"
    sql += "\n-- 1. Este script SQL foi gerado com base nas informações do seu arquivo XLSX."
    sql += "\n--    Revise os tipos de dados e tamanhos para garantir que estão corretos para PostgreSQL."
    sql += "\n-- 2. Restrições como NOT NULL, PRIMARY KEY (exceto SERIAL), UNIQUE não foram adicionadas automaticamente"
    sql += "\n--    porque não estavam no arquivo XLSX de entrada (a menos que você o modifique)."
    sql += "\n--    Você precisará adicioná-las manualmente conforme a necessidade do seu modelo."
    sql += "\n--    Se você quiser uma chave primaria SERIAL auto-gerada, adicione a linha '\"id\" SERIAL PRIMARY KEY,' após o '(' na declaração CREATE TABLE."
    sql += "\n-- 3. Nomes de colunas originais foram sanitizados (caracteres especiais removidos/substituídos, espaços por underscores, minúsculas)."
    sql += "\n--    Verifique se os nomes sanitizados são aceitáveis."

    return sql


FILE_PATH = 'dicionário_dados_educação_básica-m.xlsx' #Nome do arquivo a ser ánalisado
SHEET_NAME = None
TABLE_NAME = 'escolas_br_dados' #Nome da tabela a ser criada

COLUMN_NAME_COL_IDX = 1
COLUMN_TYPE_COL_IDX = 2
COLUMN_SIZE_COL_IDX = 3
HEADER_ROW_IDX = 1

if __name__ == "__main__":
    print(
        f"Tentando gerar script CREATE TABLE a partir do arquivo XLSX: '{FILE_PATH}'")
    print(
        f"Lendo a planilha: {'ativa' if SHEET_NAME is None else SHEET_NAME}, a partir da linha {HEADER_ROW_IDX + 1}")

    create_sql = generate_create_table_sql_from_xlsx(
        FILE_PATH,
        sheet_name=SHEET_NAME,
        table_name=TABLE_NAME,
        column_name_col_idx=COLUMN_NAME_COL_IDX,
        column_type_col_idx=COLUMN_TYPE_COL_IDX,
        column_size_col_idx=COLUMN_SIZE_COL_IDX,
        header_row_idx=HEADER_ROW_IDX
    )

    print("\n--- Script SQL CREATE TABLE ---")
    print(create_sql)
    print("\n--- Fim do Script SQL ---")

    if not create_sql.startswith("-- Erro:"):
        print("\nCopie o script SQL acima e execute-o na sua IDE de PostgreSQL.")
        print("Revise cuidadosamente os tipos de dados, tamanhos e adicione restrições (PRIMARY KEY, NOT NULL) conforme necessário.")
    else:
        print("\nNão foi possível gerar o script SQL.")
        print("Verifique o caminho do arquivo, nome da planilha, índices das colunas e se o arquivo não está vazio ou corrompido.")
